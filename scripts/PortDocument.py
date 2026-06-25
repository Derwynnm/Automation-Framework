import argparse
import re
import openpyxl
from netmiko import ConnectHandler
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import defaultdict
import logging

from dotenv import load_dotenv
from pathlib import Path
import os

env_path = Path(__file__).resolve().parents[1] / ".env"  # go up to repo root
load_dotenv(env_path)

# ----- Calling .env
username = os.getenv("NET_USERNAME")
password = os.getenv("NET_PASSWORD")
secret = os.getenv("NET_SECRET")

# Configuration

THREADS = 50  # 3 switches × up to 8 stack members per sheet = 24 max concurrent sessions
FILE_PATH = Path(__file__).resolve().parents[1] / "database" / "PortPush.xlsx"  # Excel input
SSH_TIMEOUT = 60  # Seconds before an SSH attempt times out
LOG_FILE = Path(__file__).resolve().parents[1] / "logs" / "port_description_log.txt"  # Log file name

# Setup logging
logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Worker functions – run in their own threads

def configure_port_description(switch_ip: str, port_desc_pairs: list[tuple[str, str, bool]]) -> str:
    """Connects to a Cisco IOS device, applies descriptions/shutdowns to all interfaces in one session."""

    device = {
        "device_type": "cisco_ios",
        "host": switch_ip,
        "username": username,
        "password": password,
        "secret": secret if secret else None,
        "conn_timeout": SSH_TIMEOUT,
        "verbose": False,
    }

    try:
        with ConnectHandler(**device) as net_connect:
            net_connect.enable()

            cmds = []
            skipped = []
            for port, description, shutdown in port_desc_pairs:
                if port.lower().startswith(('po', 'twe')):
                    skipped.append(port)
                    logging.info(f"Skipped {port} on {switch_ip} (Port-Channel or Twe)")
                    continue
                if shutdown:
                    cmds.extend([f"interface {port}", "no description", "shutdown"])
                    logging.info(f"Shutting down {port} on {switch_ip} (red cell)")
                elif description:
                    cmds.extend([f"interface {port}", f"description {description}"])
                else:
                    cmds.extend([f"interface {port}", "no description"])

            configured = len(port_desc_pairs) - len(skipped)
            if cmds:
                response = net_connect.send_config_set(cmds, cmd_verify=False, read_timeout=120)
                if "% Invalid input" in response or "% Incomplete" in response or "% Ambiguous" in response:
                    logging.warning(f"Potential issue on {switch_ip}: {response.strip()}")
                else:
                    logging.info(f"Configured {configured} port(s) on {switch_ip}")

        result = f"SUCCESS: {switch_ip} ({configured} ports)"
        logging.info(result)
        return result

    except Exception as exc:
        error = f"FAIL:    {switch_ip} → {exc}"
        logging.error(error)
        return error


def write_switch_config(switch_ip: str) -> str:
    """Opens a fresh SSH session to the switch and saves the running config."""

    device = {
        "device_type": "cisco_ios",
        "host": switch_ip,
        "username": username,
        "password": password,
        "secret": secret if secret else None,
        "conn_timeout": SSH_TIMEOUT,
        "verbose": False,
    }

    try:
        with ConnectHandler(**device) as net_connect:
            net_connect.enable()
            net_connect.save_config()
        result = f"WRITE:   {switch_ip}"
        logging.info(result)
        return result

    except Exception as exc:
        error = f"WRITE FAIL: {switch_ip} → {exc}"
        logging.error(error)
        return error


# Helpers

def _stack_member(port: str) -> str:
    """Extract stack member number from an interface name, e.g. 'Gi2/0/5' → '2'."""
    m = re.match(r'[A-Za-z]+(\d+)/', port)
    return m.group(1) if m else "0"

def _is_red_cell(cell) -> bool:
    """Return True if the cell has a solid red background fill."""
    fill = cell.fill
    if fill.fill_type != "solid":
        return False
    rgb = fill.fgColor.rgb.upper()  # ARGB string e.g. 'FFFF0000'
    return rgb.endswith("FF0000")

def _is_valid_ip(value: str) -> bool:
    parts = value.split(".")
    return len(parts) == 4 and all(p.isdigit() for p in parts)

def _read_sheet(ws) -> list[dict]:
    """Read one worksheet and return port rows. Handles up to 3 side-by-side switch blocks.
    Headers are in row 2; data starts at row 3. Colored label rows (no valid IP) inherit
    the IP of the first valid-IP row below them in the same block.
    """
    OFFSETS = {"IP Address": 0, "Port": 1, "Patch": 2, "Logical Patch": 3, "Device": 5, "Old Description": 6}

    rows = []
    block_starts = [cell.column for cell in ws[2] if cell.value == "IP Address"]

    for start in block_starts:
        idx = {key: start - 1 + offset for key, offset in OFFSETS.items()}

        # First pass: collect (ip_or_none, row) for every row that has a port
        raw = []
        for row in ws.iter_rows(min_row=3):
            port_val = row[idx["Port"]].value
            if not port_val:
                continue
            ip_val = row[idx["IP Address"]].value
            ip_str = str(ip_val).strip() if ip_val else ""
            ip = ip_str if _is_valid_ip(ip_str) else None
            raw.append([ip, row])

        # Back-fill: colored label rows at the top have no IP — give them
        # the IP of the first valid-IP row that follows them in the block.
        next_ip = None
        for entry in reversed(raw):
            if entry[0] is not None:
                next_ip = entry[0]
            else:
                entry[0] = next_ip

        # Second pass: build row dicts
        for ip, row in raw:
            if not ip:
                continue  # no valid IP found anywhere below — skip

            desc_cell = row[idx["Old Description"]]
            desc = "" if desc_cell.value is None else str(desc_cell.value).strip()

            def _str(v):
                return "" if v is None else str(v).strip()

            rows.append({
                "IP Address": ip,
                "Port": str(row[idx["Port"]].value).strip(),
                "Patch": _str(row[idx["Patch"]].value),
                "Logical Patch": _str(row[idx["Logical Patch"]].value),
                "Device": _str(row[idx["Device"]].value),
                "Description": desc,
                "Shutdown": _is_red_cell(desc_cell),
            })

    return rows


# Main execution

def main() -> None:
    parser = argparse.ArgumentParser(description="Push port descriptions to Cisco switches.")
    parser.add_argument("--sheet", default=None, help="Target a single sheet by name (default: all sheets)")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    sheets = [ws for ws in wb.worksheets if args.sheet is None or ws.title == args.sheet]

    if args.sheet and not sheets:
        print(f"WARNING: Sheet '{args.sheet}' not found. Available: {[ws.title for ws in wb.worksheets]}")
        return

    for ws in sheets:
        print(f"\n--- {ws.title} ---")

        rows = _read_sheet(ws)

        # Group by (ip, stack_member) — each stack member gets its own SSH session
        ip_jobs: dict[tuple[str, str], list[tuple[str, str, bool]]] = defaultdict(list)
        for row in rows:
            ip = row["IP Address"]
            desc = row["Description"]
            shutdown = row["Shutdown"]
            if not desc and row["Device"].strip().lower() == "verkada":
                desc = "Verkada"
            if desc:
                patch = row["Logical Patch"] or row["Patch"]
                if patch:
                    desc = f"{desc} ({patch})"
            for port in row["Port"].split(","):
                port = port.strip()
                if port:
                    ip_jobs[(ip, _stack_member(port))].append((port, desc, shutdown))

        if not ip_jobs:
            print("  No valid jobs — skipping.")
            continue

        # Push all port configs — one thread per stack member, up to THREADS at once
        with ThreadPoolExecutor(max_workers=min(THREADS, len(ip_jobs))) as pool:
            futures = {pool.submit(configure_port_description, ip, pairs): (ip, member)
                       for (ip, member), pairs in ip_jobs.items()}
            for future in as_completed(futures):
                print(future.result())

        # Write config once per unique switch IP (not once per stack member)
        unique_ips = sorted(set(ip for ip, _ in ip_jobs))
        print(f"  Writing config on {len(unique_ips)} switch(es)...")
        with ThreadPoolExecutor(max_workers=min(THREADS, len(unique_ips))) as pool:
            futures = {pool.submit(write_switch_config, ip): ip for ip in unique_ips}
            for future in as_completed(futures):
                print(future.result())

if __name__ == "__main__":
    main()

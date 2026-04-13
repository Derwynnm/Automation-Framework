from __future__ import annotations
import re
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from dotenv import load_dotenv
import os

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from netmiko import ConnectHandler, NetmikoAuthenticationException, NetmikoTimeoutException

# ───────────────── CONFIG ─────────────────────────────────────────────────────
EXCEL_FILE  = Path(__file__).resolve().parents[1] / "database" / "District.xlsx"
OUT_FILE    = Path(__file__).resolve().parents[1] / "docs" / "vlan_report.xlsx"
MAX_THREADS = 200
RETRIES     = 1
BACKOFF_SEC = 2
SKIP_VLANS  = {"1002", "1003", "1004", "1005"}
# -----------------------------------------------------------------------------

load_dotenv(Path(__file__).resolve().parents[1] / ".env")

USERNAME = os.getenv("NET_USERNAME")
PASSWORD = os.getenv("NET_PASSWORD")
SECRET   = os.getenv("NET_SECRET")

if not USERNAME or not PASSWORD:
    raise EnvironmentError("NET_USERNAME or NET_PASSWORD not loaded – check .env")

print(f"[INFO] Connecting as: {USERNAME}")


def parse_vlan_brief(output: str) -> tuple[list[dict], list[dict]]:
    """Return (vlans_in_use, vlans_empty) from 'show vlan brief'."""
    vlans: dict[str, dict] = {}
    current = None

    for line in output.splitlines():
        m = re.match(r"^\s*(\d{1,4})\s+(\S+)\s+\S+\s*(.*)", line)
        if m:
            vid = m.group(1)
            if vid in SKIP_VLANS:
                current = None
                continue
            current = vid
            ports_raw = m.group(3).strip()
            ports = [p.strip() for p in ports_raw.split(",") if p.strip()]
            vlans[vid] = {"id": vid, "name": m.group(2), "ports": ports}
        elif current and re.match(r"^\s{20,}", line):
            extra = line.strip()
            if extra:
                vlans[current]["ports"].extend(p.strip() for p in extra.split(",") if p.strip())

    in_use = sorted([v for v in vlans.values() if v["ports"]],     key=lambda v: int(v["id"]))
    empty  = sorted([v for v in vlans.values() if not v["ports"]], key=lambda v: int(v["id"]))
    return in_use, empty


def query_device(ip: str) -> dict | None:
    device = {
        "device_type": "cisco_ios",
        "host":        ip,
        "username":    USERNAME,
        "password":    PASSWORD,
        "secret":      SECRET or None,
        "conn_timeout": 60,
        "auth_timeout": 30,
    }

    for attempt in range(RETRIES):
        try:
            conn = ConnectHandler(**device)
            if not conn.check_enable_mode():
                conn.enable()
            output = conn.send_command("show vlan brief")
            conn.disconnect()
            in_use, empty = parse_vlan_brief(output)
            return {"in_use": in_use, "empty": empty}
        except NetmikoAuthenticationException as e:
            print(f"[AUTHFAIL] {ip} – {e}")
            return None
        except NetmikoTimeoutException:
            wait = BACKOFF_SEC * 2 ** attempt
            print(f"[TIMEOUT] {ip} – retry {attempt + 1}/{RETRIES} in {wait}s")
            time.sleep(wait)

    print(f"[FAIL] {ip} – could not connect after {RETRIES} retries")
    return None


def build_site_summary(device_results: list[dict]) -> list[dict]:
    """
    Group devices by first 3 letters of name.
    For each site+VLAN combination, report whether any device in the site has it in use.
    """
    # site_code -> vlan_id -> {"vlan_name", "active_on": [...], "empty_on": [...]}
    sites: dict[str, dict] = defaultdict(lambda: defaultdict(lambda: {"vlan_name": "", "active_on": [], "empty_on": []}))

    for r in device_results:
        name = r["name"]
        data = r["data"]
        if data is None:
            continue
        site = name[:3].upper()

        for v in data["in_use"]:
            sites[site][v["id"]]["vlan_name"] = v["name"]
            sites[site][v["id"]]["active_on"].append(name)

        for v in data["empty"]:
            sites[site][v["id"]]["vlan_name"] = v["name"]
            sites[site][v["id"]]["empty_on"].append(name)

    rows = []
    for site in sorted(sites):
        for vid in sorted(sites[site], key=int):
            entry = sites[site][vid]
            site_status = "In Use" if entry["active_on"] else "Empty"
            active_str  = ", ".join(f"Present in {d}" for d in sorted(entry["active_on"]))
            empty_str   = ", ".join(f"Not in {d}"     for d in sorted(entry["empty_on"]))
            rows.append({
                "Site":          site,
                "VLAN ID":       vid,
                "VLAN Name":     entry["vlan_name"],
                "Site Status":   site_status,
                "Active On":     active_str,
                "Empty On":      empty_str,
            })

    return rows


def style_sheet(ws, header_color: str) -> None:
    """Apply formatting to a worksheet."""
    header_fill   = PatternFill("solid", fgColor=header_color)
    in_use_fill   = PatternFill("solid", fgColor="C6EFCE")   # green
    empty_fill    = PatternFill("solid", fgColor="FFEB9C")   # yellow
    failed_fill   = PatternFill("solid", fgColor="FFC7CE")   # red
    alt_fill      = PatternFill("solid", fgColor="F2F2F2")   # light grey
    thin_border   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border
    ws.row_dimensions[1].height = 20

    status_col = next(
        (i for i, c in enumerate(ws[1], 1) if "status" in str(c.value).lower()),
        None,
    )

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        for cell in row:
            cell.border    = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.fill      = fill

        if status_col:
            status_val = str(ws.cell(row=row_idx, column=status_col).value or "")
            if status_val == "In Use":
                ws.cell(row=row_idx, column=status_col).fill = in_use_fill
            elif status_val == "Empty":
                ws.cell(row=row_idx, column=status_col).fill = empty_fill
            elif "FAIL" in status_val:
                ws.cell(row=row_idx, column=status_col).fill = failed_fill

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    ws.freeze_panes = "A2"


def main() -> None:
    if not EXCEL_FILE.is_file():
        raise FileNotFoundError(EXCEL_FILE)

    df = pd.read_excel(EXCEL_FILE)

    ip_col   = next((c for c in df.columns if "ip"   in c.lower()), None)
    name_col = next((c for c in df.columns if c.lower() in ("name", "site", "hostname", "device name", "device")), None)

    if not ip_col:
        raise ValueError(f"No IP column found. Available columns: {list(df.columns)}")

    print(f"IP column: '{ip_col}'  |  Name column: '{name_col or 'none – using IP'}'")
    print(f"Loaded {len(df)} rows – starting with up to {MAX_THREADS} threads\n")

    valid_rows = [
        (idx, str(row[ip_col]).strip(), str(row[name_col]).strip() if name_col else str(row[ip_col]).strip())
        for idx, row in df.iterrows()
        if str(row[ip_col]).strip() not in ("", "nan", "None")
    ]

    results: dict[int, dict] = {}
    with ThreadPoolExecutor(max_workers=MAX_THREADS) as pool:
        futures = {
            pool.submit(query_device, ip): (idx, name, ip)
            for idx, ip, name in valid_rows
        }
        for fut in as_completed(futures):
            idx, name, ip = futures[fut]
            try:
                data = fut.result()
            except Exception as exc:
                print(f"[EXC] {ip}: {exc}")
                data = None

            if data:
                print(f"[OK] {ip} ({name}) – {len(data['in_use'])} in use / {len(data['empty'])} empty")
            results[idx] = {"name": name, "ip": ip, "data": data}

    # ── Sheet 1: per-device detail ────────────────────────────────────────────
    device_rows = []
    device_results = []
    for idx in sorted(results):
        r = results[idx]
        device_results.append(r)
        name, ip, data = r["name"], r["ip"], r["data"]

        if data is None:
            device_rows.append({"Name": name, "IP Address": ip, "VLAN ID": "",
                                 "VLAN Name": "", "Status": "CONNECTION FAILED", "Ports": ""})
            continue

        for v in data["in_use"]:
            device_rows.append({"Name": name, "IP Address": ip, "VLAN ID": v["id"],
                                 "VLAN Name": v["name"], "Status": "In Use", "Ports": ", ".join(v["ports"])})
        for v in data["empty"]:
            device_rows.append({"Name": name, "IP Address": ip, "VLAN ID": v["id"],
                                 "VLAN Name": v["name"], "Status": "Empty", "Ports": ""})

    # ── Sheet 2: site summary ─────────────────────────────────────────────────
    site_rows = build_site_summary(device_results)

    # ── Write Excel ───────────────────────────────────────────────────────────
    OUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUT_FILE, engine="openpyxl") as writer:
        pd.DataFrame(device_rows, columns=["Name", "IP Address", "VLAN ID", "VLAN Name", "Status", "Ports"]
                     ).to_excel(writer, sheet_name="Device Detail", index=False)

        pd.DataFrame(site_rows, columns=["Site", "VLAN ID", "VLAN Name", "Site Status", "Active On", "Empty On"]
                     ).to_excel(writer, sheet_name="Site Summary", index=False)

        style_sheet(writer.sheets["Device Detail"], header_color="2E75B6")
        style_sheet(writer.sheets["Site Summary"],  header_color="375623")

    print(f"\nReport written to {OUT_FILE}")
    print(f"  Device Detail : {len(device_rows)} rows")
    print(f"  Site Summary  : {len(site_rows)} rows across {len({r['Site'] for r in site_rows})} sites")


if __name__ == "__main__":
    main()
